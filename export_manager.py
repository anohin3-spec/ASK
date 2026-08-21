"""
Модуль для экспорта данных в Excel
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportManager:
    def __init__(self, db):
        self.db = db
    
    def export_all(self, file_path):
        """Экспорт всех данных в Excel"""
        wb = Workbook()
        
        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Создаем листы
        self._create_equipment_sheet(wb)
        self._create_drivers_sheet(wb)
        self._create_maintenance_sheet(wb)
        self._create_issues_sheet(wb)
        
        # Сохраняем файл
        wb.save(file_path)
    
    def export_equipment(self, file_path):
        """Экспорт только техники"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self._create_equipment_sheet(wb)
        wb.save(file_path)
    
    def export_maintenance_history(self, file_path):
        """Экспорт истории ТО"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self._create_maintenance_sheet(wb)
        wb.save(file_path)
    
    def export_issues(self, file_path):
        """Экспорт неисправностей"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self._create_issues_sheet(wb)
        wb.save(file_path)
    
    def _create_equipment_sheet(self, wb):
        """Создание листа с техникой"""
        ws = wb.create_sheet("Техника")
        
        # Заголовки
        headers = [
            'ID', 'Наименование техники', 'VIN', 'СТС (текст)', 'Номер', 'Тип учета',
            'Последнее ТО', 'Текущее значение', 'Интервал ТО (лето)',
            'Интервал ТО (зима)', 'Следующее ТО', 'До ТО осталось',
            'Водители', 'Ситуация', 'Сервис', 'Страховка', 'Пропуск МКАД'
        ]
        
        # Стили заголовка
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        
        # Получаем данные
        equipment_list = self.db.get_all_equipment()
        current_month = datetime.now().month
        # Зима: ноябрь-февраль (11,12,1,2), Лето: март-октябрь (3-10)
        is_winter = current_month in [11, 12, 1, 2]
        
        # Записываем данные
        for row_num, eq in enumerate(equipment_list, 2):
            # Получаем водителей
            drivers = self.db.get_equipment_drivers(eq['id'])
            drivers_str = ', '.join([d['name'] for d in drivers])
            
            # Расчет следующего ТО
            interval = eq['maintenance_interval_winter'] if is_winter else eq['maintenance_interval_summer']
            next_maintenance = eq['last_maintenance'] + interval
            remaining = next_maintenance - eq['current_value']
            
            measurement_type = 'Пробег' if eq['measurement_type'] == 'mileage' else 'Моточасы'
            try:
                sts_certificate = eq['sts_certificate'] or ''
            except (KeyError, IndexError, TypeError):
                sts_certificate = ''
            
            data = [
                eq['id'],
                eq['name'],
                eq['sts_pts'],
                sts_certificate,
                eq['reg_number'],
                measurement_type,
                eq['last_maintenance'],
                eq['current_value'],
                eq['maintenance_interval_summer'],
                eq['maintenance_interval_winter'],
                next_maintenance,
                remaining,
                drivers_str,
                eq['situation'],
                eq['service'],
                eq['insurance_date'],
                eq['mkad_pass_date'] if eq['mkad_pass_date'] else ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Цвет строки в зависимости от статуса ТО
                if remaining <= 0:
                    cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                elif remaining <= interval * 0.2:
                    cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        # Автоширина столбцов
        self._autofit_columns(ws)
        
        # Закрепление первой строки
        ws.freeze_panes = 'A2'
    
    def _create_drivers_sheet(self, wb):
        """Создание листа с водителями"""
        ws = wb.create_sheet("Водители")
        
        headers = ['ID', 'Имя', 'Телефон', 'Закрепленная техника']
        
        # Стили
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        
        # Данные
        drivers_list = self.db.get_all_drivers()
        
        for row_num, driver in enumerate(drivers_list, 2):
            equipment_list = self.db.get_driver_equipment(driver['id'])
            equipment_str = ', '.join([f"{eq['name']} ({eq['reg_number']})" for eq in equipment_list])
            
            data = [
                driver['id'],
                driver['name'],
                driver['phone'],
                equipment_str
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
        
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_maintenance_sheet(self, wb):
        """Создание листа с историей ТО"""
        ws = wb.create_sheet("История ТО")
        
        headers = [
            'ID', 'Техника', 'Номер', 'Пробег/моточасы при ТО',
            'Дата ТО', 'Комментарий', 'Путь к счету'
        ]
        
        # Стили
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        
        # Данные
        maintenance_list = self.db.get_all_maintenance_history()
        
        for row_num, maint in enumerate(maintenance_list, 2):
            date_str = datetime.fromisoformat(maint['maintenance_date']).strftime('%d.%m.%Y %H:%M')
            
            data = [
                maint['id'],
                maint['equipment_name'],
                maint['reg_number'],
                maint['maintenance_value'],
                date_str,
                maint['comment'],
                maint['invoice_path']
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
        
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_issues_sheet(self, wb):
        """Создание листа с неисправностями"""
        ws = wb.create_sheet("Неисправности")
        
        headers = [
            'ID', 'Техника', 'Номер', 'Водитель', 'Телефон водителя',
            'Описание', 'Статус', 'Дата сообщения', 'Дата решения',
            'Комментарий к решению'
        ]
        
        # Стили
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border
        
        # Данные
        issues_list = self.db.get_all_issues()
        
        for row_num, issue in enumerate(issues_list, 2):
            reported_date = datetime.fromisoformat(issue['reported_date']).strftime('%d.%m.%Y %H:%M')
            resolved_date = ''
            if issue['resolved_date']:
                resolved_date = datetime.fromisoformat(issue['resolved_date']).strftime('%d.%m.%Y %H:%M')
            
            status = 'Открыто' if issue['status'] == 'open' else 'Закрыто'
            
            data = [
                issue['id'],
                issue['equipment_name'],
                issue['reg_number'],
                issue['driver_name'] or '-',
                issue['driver_phone'] or '-',
                issue['description'],
                status,
                reported_date,
                resolved_date,
                issue['resolution_comment'] or ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Цвет строки в зависимости от статуса
                if issue['status'] == 'open':
                    cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        self._autofit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _autofit_columns(self, ws):
        """Автоматическая подгонка ширины столбцов"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
